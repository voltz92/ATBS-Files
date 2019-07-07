import openpyxl
import os

# first check whether the working directory is in the excel files location or not, so that all relative path commands work
if os.getcwd() != 'D:\\Files\\Git Repos\\ATBS-Files\\Office Files\\Excel-Files':
    print('Switching to Excel-File directory')
    os.chdir('D:\\Files\\Git Repos\\ATBS-Files\\Office Files\\Excel-Files')

# filename = input('Please enter the filename (with format, e.g filename.xlsx) >>> ')
# this wud be used to obtain filename from the user

# opening the workbook
workbook = openpyxl.load_workbook('data.xlsx')
# Creates a workbook object

# can use sheet = workbook.get_sheet_by_name('Sheet1') to get a specific sheet
# if sheet names unknown

print('This workbook contains the following sheets: ')
for sheet in workbook.sheetnames:    # the attribute workbook.sheetnames returns a list of strings of sheetNames in the excel file
    print(' > ' + sheet)
workbook.active = workbook['Sheet1']   # changing the active workbook

sheet = workbook.active    #assigning the active worksheet to a variable to access the attributes
print(str(sheet['A1'].value))    #accessing the value of a specific cell
if sheet['D1'].value == None:
    print('This cell doesnt contain any data')

print(sheet.max_column)
print(sheet.max_row)

# we can assign a cell to a var ::
c = sheet['B1']
# available methods for this object :: 
print(c.value)
print(c.row)
print(c.column)
print(c.coordinate)

# we can also access cells using the sheet.cell() method
# we can iterate using a loop using this method to access multiple cells
for i in range(1,sheet.max_row):
    for j in range(1,sheet.max_column):
        if sheet.cell(row = i, column = j).value == None:
            continue    
        print(sheet.cell(row = i, column = j).value)


