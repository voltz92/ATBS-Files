import openpyxl
import os

# change directory
os.chdir('D:\\Files\\Git Repos\\ATBS-Files\\Office Files\\Excel-Files')

# creating a new Workbook object
wb = openpyxl.Workbook()

# by default created with a sheet called : 
print(wb.sheetnames)

sheet = wb['Sheet']
# writing random data
for i in range(1,11):  # here starting range from 1 as the cell's row/column index starts at 1
    sheet.cell(row = i, column = 1).value = i
sheet['B2'] = 'Sup'

# to save this object as an xlsx file
wb.save('writing_to_excel_with_python.xlsx')

# creating a new sheet

sheet2 = wb.create_sheet('Created with openpyxl')  # appends the new sheet to the end. to create a sheet a specific position just add the index = postion arg to the method
print(wb.sheetnames)
sheet2.title = 'Changing title with python'  # changing the title by using the attribute of the sheet2 object
wb.save('to_prevent_overwriting.xlsx')  # usually save with a different filename to prevent overwriting data due to bugs

